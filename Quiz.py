import tkinter as tk
from tkinter import ttk
import random
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
# Create Excel file if it doesn't exist
if not os.path.exists("quiz_progress.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quiz Progress"
    sheet.append(["Name", "Timestamp", "Score (%)", "Total Questions", "Attempt", "Difficulty", "Course", "Time Taken (s)", "Feedback"])
    workbook.save("quiz_progress.xlsx")
question_bank = { # Centralized question bank
    "elementary": {
        "math": [
            {"question": "What is 5 + 3?", "options": ["6", "7", "8", "9"], "answer": "8"},
            {"question": "What is 10 - 4?", "options": ["5", "6", "7", "8"], "answer": "6"},
            {"question": "What is 2 x 6?", "options": ["10", "12", "14", "16"], "answer": "12"},
            {"question": "What is 15 / 3?", "options": ["4", "5", "6", "7"], "answer": "5"},
            {"question": "What is 9 + 1?", "options": ["8", "9", "10", "11"], "answer": "10"},
        ],
        "science": [
            {"question": "What planet is known as the Red Planet?", "options": ["Earth", "Mars", "Jupiter", "Venus"], "answer": "Mars"},
            {"question": "What is the boiling point of water?", "options": ["50°C", "100°C", "150°C", "200°C"], "answer": "100°C"},
            {"question": "What gas do plants need for photosynthesis?", "options": ["Oxygen", "Carbon Dioxide", "Nitrogen", "Hydrogen"], "answer": "Carbon Dioxide"},
            {"question": "What is H2O?", "options": ["Water", "Oxygen", "Hydrogen", "Salt"], "answer": "Water"},
            {"question": "What is the largest organ in the human body?", "options": ["Heart", "Liver", "Skin", "Lungs"], "answer": "Skin"},
        ],
        "history": [
            {"question": "Who was the first President of the United States?", "options": ["Abraham Lincoln", "George Washington", "Thomas Jefferson", "John Adams"], "answer": "George Washington"},
            {"question": "What year did Christopher Columbus discover America?", "options": ["1492", "1776", "1620", "1800"], "answer": "1492"},
            {"question": "What is the Liberty Bell a symbol of?", "options": ["Freedom", "Wealth", "Power", "Peace"], "answer": "Freedom"},
            {"question": "Who wrote the Declaration of Independence?", "options": ["Benjamin Franklin", "George Washington", "Thomas Jefferson", "John Hancock"], "answer": "Thomas Jefferson"},
            {"question": "What country gifted the Statue of Liberty to the United States?", "options": ["England", "France", "Canada", "Spain"], "answer": "France"},
        ],
        "reading": [
            {"question": "What is the main character in a story called?", "options": ["Antagonist", "Protagonist", "Narrator", "Villain"], "answer": "Protagonist"},
            {"question": "Which of these is a fiction story?", "options": ["A story about talking animals", "A biography", "A science textbook", "A history book"], "answer": "A story about talking animals"},
            {"question": "What punctuation mark is used at the end of a question?", "options": ["Period", "Exclamation Point", "Question Mark", "Comma"], "answer": "Question Mark"},
            {"question": "What do we call the person who writes a book?", "options": ["Author", "Editor", "Illustrator", "Publisher"], "answer": "Author"},
            {"question": "Which of these words is a noun?", "options": ["Run", "Happy", "Dog", "Blue"], "answer": "Dog"},
        ],
        "music": [
            {"question": "How many notes are in a musical scale?", "options": ["5", "7", "8", "12"], "answer": "8"},
            {"question": "Which family of instruments does a trumpet belong to?", "options": ["Percussion", "Strings", "Brass", "Woodwind"], "answer": "Brass"},
            {"question": "What does a quarter note represent in music?", "options": ["Half a beat", "One beat", "Two beats", "Four beats"], "answer": "One beat"},
            {"question": "What do you call the person who leads an orchestra?", "options": ["Composer", "Performer", "Conductor", "Instrumentalist"], "answer": "Conductor"},
            {"question": "Which instrument has black and white keys?", "options": ["Guitar", "Piano", "Violin", "Flute"], "answer": "Piano"},
        ],
        "spanish": [
            {"question": "What does 'Hola' mean in English?", "options": ["Goodbye", "Please", "Hello", "Thank you"], "answer": "Hello"},
            {"question": "What color is 'Rojo'?", "options": ["Blue", "Red", "Green", "Yellow"], "answer": "Red"},
            {"question": "How do you say 'Thank you' in Spanish?", "options": ["Gracias", "Hola", "Adiós", "Por favor"], "answer": "Gracias"},
            {"question": "Which day of the week is 'Lunes'?", "options": ["Monday", "Tuesday", "Wednesday", "Sunday"], "answer": "Monday"},
            {"question": "What is the Spanish word for 'House'?", "options": ["Casa", "Perro", "Gato", "Libro"], "answer": "Casa"},
        ],
    },
    "middle": {
        "math": [
            {"question": "What is the value of 7^2?", "options": ["14", "49", "21", "81"], "answer": "49"},
            {"question": "Solve for x: 5x + 3 = 18.", "options": ["2", "3", "4", "5"], "answer": "3"},
            {"question": "What is the area of a triangle with a base of 10 cm and a height of 5 cm?", "options": ["25 cm²", "50 cm²", "15 cm²", "20 cm²"], "answer": "25 cm²"},
            {"question": "What is the greatest common factor (GCF) of 36 and 48?", "options": ["6", "8", "12", "18"], "answer": "12"},
            {"question": "What is the value of 3/4 + 5/8?", "options": ["1/8", "7/8", "11/8", "5/4"], "answer": "11/8"},
        ],
        "science": [
            {"question": "What is the process by which plants make their own food?", "options": ["Respiration", "Photosynthesis", "Fermentation", "Digestion"], "answer": "Photosynthesis"},
            {"question": "What is the most abundant gas in Earth's atmosphere?", "options": ["Oxygen", "Nitrogen", "Carbon Dioxide", "Hydrogen"], "answer": "Nitrogen"},
            {"question": "What organelle is known as the 'powerhouse of the cell'?", "options": ["Nucleus", "Chloroplast", "Mitochondria", "Ribosome"], "answer": "Mitochondria"},
            {"question": "What is the chemical formula for table salt?", "options": ["H2O", "CO2", "NaCl", "KCl"], "answer": "NaCl"},
            {"question": "What is the force that pulls objects toward the center of the Earth?", "options": ["Magnetism", "Gravity", "Friction", "Inertia"], "answer": "Gravity"},
        ],
        "history": [
            {"question": "What war was fought between the North and South regions of the United States?", "options": ["Revolutionary War", "Civil War", "War of 1812", "World War I"], "answer": "Civil War"},
            {"question": "Who was the leader of the Confederate States during the Civil War?", "options": ["Abraham Lincoln", "Jefferson Davis", "Ulysses S. Grant", "Robert E. Lee"], "answer": "Jefferson Davis"},
            {"question": "What was the name of the ship the Pilgrims traveled on to America?", "options": ["Mayflower", "Santa Maria", "Titanic", "Nina"], "answer": "Mayflower"},
            {"question": "What was the name of the series of events that ended feudalism and introduced the Renaissance?", "options": ["The Dark Ages", "The Middle Ages", "The Crusades", "The Industrial Revolution"], "answer": "The Crusades"},
            {"question": "Who was known as the 'Father of the Constitution'?", "options": ["George Washington", "James Madison", "Thomas Jefferson", "Benjamin Franklin"], "answer": "James Madison"},
        ],
        "reading": [
            {"question": "What type of literature features imaginary events and characters?", "options": ["Nonfiction", "Fiction", "Biography", "Poetry"], "answer": "Fiction"},
            {"question": "What is the central idea of a story called?", "options": ["Plot", "Theme", "Characterization", "Setting"], "answer": "Theme"},
            {"question": "What is it called when words imitate sounds?", "options": ["Simile", "Metaphor", "Onomatopoeia", "Alliteration"], "answer": "Onomatopoeia"},
            {"question": "Which point of view uses 'I' or 'we'?", "options": ["First Person", "Second Person", "Third Person", "Omniscient"], "answer": "First Person"},
            {"question": "What is the climax of a story?", "options": ["The introduction of the characters", "The highest point of tension", "The resolution of the plot", "The description of the setting"], "answer": "The highest point of tension"},
        ],
        "music": [
            {"question": "What is the term for the speed of music?", "options": ["Melody", "Rhythm", "Tempo", "Harmony"], "answer": "Tempo"},
            {"question": "What are the lines and spaces on which music is written called?", "options": ["Scale", "Staff", "Ledger", "Bar"], "answer": "Staff"},
            {"question": "What is a group of sharps or flats at the beginning of a staff called?", "options": ["Key Signature", "Time Signature", "Measure", "Bar Line"], "answer": "Key Signature"},
            {"question": "What does a whole note represent?", "options": ["1 beat", "2 beats", "4 beats", "8 beats"], "answer": "4 beats"},
            {"question": "Which instrument family does a flute belong to?", "options": ["Brass", "Percussion", "String", "Woodwind"], "answer": "Woodwind"},
        ],
        "spanish": [
            {"question": "What is the Spanish word for 'Friend'?", "options": ["Casa", "Amigo", "Perro", "Libro"], "answer": "Amigo"},
            {"question": "What is the correct translation for 'Good morning'?", "options": ["Buenas noches", "Buenas tardes", "Buenos días", "Hola"], "answer": "Buenos días"},
            {"question": "How do you say 'I like' in Spanish?", "options": ["Yo quiero", "Me gusta", "Yo tengo", "Yo soy"], "answer": "Me gusta"},
            {"question": "What is the Spanish word for 'School'?", "options": ["Escuela", "Trabajo", "Clase", "Universidad"], "answer": "Escuela"},
            {"question": "What is the plural of 'libro' (book) in Spanish?", "options": ["Libros", "Libras", "Librons", "Libras"], "answer": "Libros"},
        ],
    },
    "high": {
        "math": [
            {"question": "Solve for x: 3x - 7 = 11.", "options": ["2", "6", "4", "5"], "answer": "6"},
            {"question": "What is the derivative of f(x) = x^2 + 3x?", "options": ["x + 3", "2x + 3", "2x", "x^2"], "answer": "2x + 3"},
            {"question": "What is the slope of the line 2x + 3y = 6?", "options": ["-2/3", "2/3", "-3/2", "3/2"], "answer": "-2/3"},
            {"question": "What is the quadratic formula used to solve ax^2 + bx + c = 0?", "options": ["x = (-b ± √(b^2 - 4ac)) / 2a", "x = -b + 4ac / 2a", "x = -b ± 4ac / a", "x = -b^2 + √(b^2 - 4ac) / a"], "answer": "x = (-b ± √(b^2 - 4ac)) / 2a"},
            {"question": "If sin(30°) = x, what is x?", "options": ["1/2", "√2/2", "√3/2", "1"], "answer": "1/2"},
        ],
        "science": [
            {"question": "What is Newton's third law of motion?", "options": ["Every action has an equal and opposite reaction.", "An object in motion stays in motion unless acted upon by an external force.", "Force equals mass times acceleration.", "Energy cannot be created or destroyed."], "answer": "Every action has an equal and opposite reaction."},
            {"question": "What is the chemical symbol for gold?", "options": ["Au", "Ag", "Gd", "Go"], "answer": "Au"},
            {"question": "Which part of the cell controls its activities?", "options": ["Mitochondria", "Nucleus", "Ribosome", "Cytoplasm"], "answer": "Nucleus"},
            {"question": "What is the atomic number of carbon?", "options": ["6", "8", "12", "14"], "answer": "6"},
            {"question": "What is the most basic unit of life?", "options": ["Atom", "Molecule", "Cell", "Organ"], "answer": "Cell"},
        ],
        "history": [
            {"question": "Who was the president during the Great Depression and World War II?", "options": ["Franklin D. Roosevelt", "Harry S. Truman", "Herbert Hoover", "Dwight D. Eisenhower"], "answer": "Franklin D. Roosevelt"},
            {"question": "What year did the Berlin Wall fall?", "options": ["1985", "1989", "1991", "1995"], "answer": "1989"},
            {"question": "What ancient civilization invented democracy?", "options": ["Romans", "Greeks", "Egyptians", "Mesopotamians"], "answer": "Greeks"},
            {"question": "What treaty ended World War I?", "options": ["Treaty of Versailles", "Treaty of Paris", "Treaty of Tordesillas", "Treaty of Ghent"], "answer": "Treaty of Versailles"},
            {"question": "Who was the leader of the Soviet Union during World War II?", "options": ["Vladimir Lenin", "Joseph Stalin", "Nikita Khrushchev", "Mikhail Gorbachev"], "answer": "Joseph Stalin"},
        ],
        "reading": [
            {"question": "What type of conflict involves a character struggling with their own mind?", "options": ["Man vs. Man", "Man vs. Nature", "Man vs. Self", "Man vs. Society"], "answer": "Man vs. Self"},
            {"question": "What is the purpose of a thesis statement in an essay?", "options": ["To provide evidence", "To summarize the main argument", "To list sources", "To explain the setting"], "answer": "To summarize the main argument"},
            {"question": "What does the term 'foreshadowing' mean in literature?", "options": ["A hint about future events", "A detailed description of the setting", "A conflict resolution technique", "A character's backstory"], "answer": "A hint about future events"},
            {"question": "Which literary device compares two things using 'like' or 'as'?", "options": ["Metaphor", "Simile", "Hyperbole", "Personification"], "answer": "Simile"},
            {"question": "What is a protagonist?", "options": ["The main character of a story", "The villain of a story", "The narrator of a story", "The setting of a story"], "answer": "The main character of a story"},
        ],
        "music": [
            {"question": "What does the term 'crescendo' mean in music?", "options": ["Gradually get louder", "Gradually get softer", "Repeat a section", "Play very fast"], "answer": "Gradually get louder"},
            {"question": "Which key has no sharps or flats?", "options": ["C Major", "G Major", "D Major", "F Major"], "answer": "C Major"},
            {"question": "What is the term for the highness or lowness of a sound?", "options": ["Tempo", "Dynamics", "Pitch", "Timbre"], "answer": "Pitch"},
            {"question": "What symbol indicates the end of a piece of music?", "options": ["Double bar line", "Repeat sign", "Fermata", "Clef"], "answer": "Double bar line"},
            {"question": "Which composer wrote the 'Moonlight Sonata'?", "options": ["Mozart", "Beethoven", "Bach", "Chopin"], "answer": "Beethoven"},
        ],
        "spanish": [
            {"question": "What is the Spanish verb for 'to eat'?", "options": ["Comer", "Beber", "Correr", "Vivir"], "answer": "Comer"},
            {"question": "What is the correct translation for 'How are you?' (formal)?", "options": ["¿Cómo estás?", "¿Qué tal?", "¿Cómo está usted?", "¿Dónde estás?"], "answer": "¿Cómo está usted?"},
            {"question": "What is the Spanish word for 'Bookstore'?", "options": ["Biblioteca", "Libro", "Librería", "Clase"], "answer": "Librería"},
            {"question": "Which word is feminine in Spanish?", "options": ["Niño", "Hombre", "Mujer", "Gato"], "answer": "Mujer"},
            {"question": "What is the Spanish translation for 'We are'?", "options": ["Soy", "Somos", "Está", "Estoy"], "answer": "Somos"},
        ],
    },
    "advanced": {
        "math": [
            {"question": "What is the integral of sin(x) with respect to x?", "options": ["-cos(x) + C", "cos(x) + C", "-sin(x) + C", "tan(x) + C"], "answer": "-cos(x) + C"},
            {"question": "Solve for x: log2(x) = 5.", "options": ["10", "25", "32", "5"], "answer": "32"},
            {"question": "What is the determinant of the matrix [[2, 3], [1, 4]]?", "options": ["5", "6", "7", "8"], "answer": "5"},
            {"question": "What is the derivative of ln(x)?", "options": ["x", "1/x", "x^2", "ln(x^2)"], "answer": "1/x"},
            {"question": "What is the value of e^0?", "options": ["0", "1", "e", "Undefined"], "answer": "1"},
        ],
        "science": [
            {"question": "What is the second law of thermodynamics?", "options": ["Energy cannot be created or destroyed.", "Entropy of an isolated system always increases.", "Force equals mass times acceleration.", "For every action, there is an equal and opposite reaction."], "answer": "Entropy of an isolated system always increases."},
            {"question": "What is the pH of pure water?", "options": ["0", "7", "14", "5"], "answer": "7"},
            {"question": "What is the charge of an electron?", "options": ["Positive", "Negative", "Neutral", "It depends on the element"], "answer": "Negative"},
            {"question": "What type of bond forms between water molecules?", "options": ["Covalent", "Ionic", "Hydrogen", "Metallic"], "answer": "Hydrogen"},
            {"question": "What is the primary function of ribosomes?", "options": ["DNA replication", "Protein synthesis", "Energy production", "Lipid storage"], "answer": "Protein synthesis"},
        ],
        "history": [
            {"question": "Which empire was known as the 'Land of the Rising Sun'?", "options": ["China", "Japan", "Mongolia", "Korea"], "answer": "Japan"},
            {"question": "What year did the United Nations form?", "options": ["1942", "1945", "1950", "1939"], "answer": "1945"},
            {"question": "Who was the last ruler of the Roman Empire?", "options": ["Julius Caesar", "Romulus Augustulus", "Augustus Caesar", "Constantine"], "answer": "Romulus Augustulus"},
            {"question": "What was the main cause of the Cold War?", "options": ["Competition for nuclear weapons", "Ideological conflict between capitalism and communism", "Disputes over oil resources", "Border conflicts in Europe"], "answer": "Ideological conflict between capitalism and communism"},
            {"question": "What event started World War I?", "options": ["The assassination of Archduke Franz Ferdinand", "Germany invading Poland", "The sinking of the Lusitania", "The bombing of Pearl Harbor"], "answer": "The assassination of Archduke Franz Ferdinand"},
        ],
        "reading": [
            {"question": "What is an allegory?", "options": ["A direct comparison between two things", "A story with a hidden meaning or moral", "The repetition of consonant sounds", "A humorous play on words"], "answer": "A story with a hidden meaning or moral"},
            {"question": "What does the term 'diction' refer to in literature?", "options": ["Sentence structure", "Word choice", "Character development", "Plot development"], "answer": "Word choice"},
            {"question": "What is the purpose of a soliloquy in drama?", "options": ["To develop the setting", "To reveal a character's inner thoughts", "To resolve the conflict", "To create suspense"], "answer": "To reveal a character's inner thoughts"},
            {"question": "What type of irony occurs when the audience knows something the characters do not?", "options": ["Situational irony", "Verbal irony", "Dramatic irony", "Cosmic irony"], "answer": "Dramatic irony"},
            {"question": "What is the term for a recurring theme or idea in a literary work?", "options": ["Symbol", "Motif", "Allusion", "Paradox"], "answer": "Motif"},
        ],
        "music": [
            {"question": "What is the dominant chord in the key of C major?", "options": ["G major", "D minor", "F major", "A minor"], "answer": "G major"},
            {"question": "What is the purpose of a fermata in music?", "options": ["To indicate a rest", "To hold a note longer than usual", "To repeat a section", "To play softly"], "answer": "To hold a note longer than usual"},
            {"question": "Who composed 'The Four Seasons'?", "options": ["Beethoven", "Vivaldi", "Mozart", "Bach"], "answer": "Vivaldi"},
            {"question": "What does the term 'fortissimo' mean?", "options": ["Very loud", "Very soft", "Gradually louder", "Gradually softer"], "answer": "Very loud"},
            {"question": "What is the name of the sequence of whole and half steps in a major scale?", "options": ["Whole-half pattern", "Major pattern", "W-W-H-W-W-W-H", "H-H-W-W-H-H"], "answer": "W-W-H-W-W-W-H"},
        ],
        "spanish": [
            {"question": "What is the future tense of 'yo comer'?", "options": ["Comeré", "Comí", "Como", "Comía"], "answer": "Comeré"},
            {"question": "How do you say 'They are studying' in Spanish?", "options": ["Estudian", "Están estudiando", "Estudiaron", "Estudiarán"], "answer": "Están estudiando"},
            {"question": "Which word means 'to learn'?", "options": ["Aprender", "Enseñar", "Escuchar", "Hablar"], "answer": "Aprender"},
            {"question": "What is the Spanish word for 'bottle'?", "options": ["Bolsa", "Botella", "Basura", "Barco"], "answer": "Botella"},
            {"question": "What does '¿Cuánto cuesta?' mean?", "options": ["Where is it?", "How are you?", "How much does it cost?", "What is it?"], "answer": "How much does it cost?"},
        ],
    },
}
# Global variables
filtered_questions = []
current_question_index = 0
score = 0
wrong_questions = []
time_remaining = 0
quiz_timer = 300
# Save progress to Excel
def save_progress(name, score, total_questions, time_taken, difficulty, course, feedback):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    workbook = load_workbook("quiz_progress.xlsx")
    sheet = workbook["Quiz Progress"]
    attempt_count = 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[5] == difficulty and row[6] == course:
            attempt_count += 1
    percentage = (score / total_questions) * 100
    sheet.append([
        name,
        timestamp,
        f"{percentage:.2f}%",
        total_questions,
        attempt_count,
        difficulty,
        course,
        time_taken,
        feedback
    ])
    workbook.save("quiz_progress.xlsx")
# Start timer
def start_timer(duration):
    global time_remaining, quiz_timer
    time_remaining = duration
    quiz_timer = duration
    update_timer()
# Update timer
def update_timer():
    global time_remaining
    if time_remaining > 0:
        time_label.config(text=f"Time Left: {time_remaining}s")
        time_remaining -= 1
        root.after(1000, update_timer)
    else:
        time_label.config(text="Time's up!")
        end_quiz()
# Filter questions
def filter_questions(difficulty, course=None):
    global filtered_questions
    filtered_questions = []
    if course and course != "difficulty test":
        filtered_questions = question_bank[difficulty].get(course, [])
    else:
        for questions in question_bank[difficulty].values():
            filtered_questions.extend(questions)
        random.shuffle(filtered_questions)
# Load a question
def load_question():
    global current_question_index
    if current_question_index < len(filtered_questions):
        question_data = filtered_questions[current_question_index]
        question_label.config(text=question_data["question"])
        for idx, option in enumerate(question_data["options"]):
            option_buttons[idx].config(text=option, state="normal", bg="lightgrey")
        result_label.config(text="")
        next_button.config(state="disabled")
    else:
        end_quiz()
# Check the answer
def check_answer(selected_option):
    global score, wrong_questions
    for button in option_buttons:
        button.config(state="disabled")
    question_data = filtered_questions[current_question_index]
    if selected_option == question_data["answer"]:
        result_label.config(text="Correct!", fg="green")
        score += 1
    else:
        result_label.config(text=f"Wrong! Correct answer: {question_data['answer']}", fg="red")
        wrong_questions.append(question_data["question"])
    next_button.config(state="normal")
# Move to the next question
def next_question():
    global current_question_index
    current_question_index += 1
    load_question()
# End the quiz
def end_quiz():
    global score, wrong_questions
    feedback = f"Your score: {score}/{len(filtered_questions)}\n"
    if wrong_questions:
        feedback += "Review these questions:\n" + "\n".join(f"- {q}" for q in wrong_questions)
    else:
        feedback += "Great job! Perfect score!"
    time_taken = quiz_timer - time_remaining
    for widget in root.winfo_children():
        widget.destroy()
    tk.Label(root, text="Enter Your Name", font=("Arial", 14)).pack(pady=5)
    name_var = tk.StringVar()
    tk.Entry(root, textvariable=name_var, font=("Arial", 14)).pack(pady=5)
    def save_and_exit():
        name = name_var.get().strip()
        course = course_var.get()
        save_progress(name, score, len(filtered_questions), time_taken, difficulty_var.get(), course, feedback)
        start_screen()
    tk.Button(root, text="Submit", font=("Arial", 14), command=save_and_exit).pack(pady=20)
# Start screen
def start_screen():
    global current_question_index, score, wrong_questions
    current_question_index = 0
    score = 0
    wrong_questions = []
    for widget in root.winfo_children():
        widget.destroy()
    tk.Label(root, text="Select Difficulty", font=("Arial", 16)).pack(pady=10)
    global difficulty_var, course_var
    difficulty_var = tk.StringVar(value="elementary")
    ttk.Combobox(root, textvariable=difficulty_var, values=list(question_bank.keys())).pack(pady=5)
    tk.Label(root, text="Select Course", font=("Arial", 16)).pack(pady=10)
    course_var = tk.StringVar(value="math")
    ttk.Combobox(root, textvariable=course_var, values=list(question_bank["elementary"].keys()) + ["difficulty test"]).pack(pady=5)
    def start_quiz():
        course = course_var.get()
        filter_questions(difficulty_var.get(), course)
        if filtered_questions:
            duration = 180 if course == "difficulty test" else 30
            load_question_screen(duration)
        else:
            tk.Label(root, text="No questions found for the selected criteria.", fg="red", font=("Arial", 12)).pack(pady=10)
    tk.Button(root, text="Start Quiz", font=("Arial", 14), command=start_quiz).pack(pady=20)
# Load quiz screen
def load_question_screen(duration):
    for widget in root.winfo_children():
        widget.destroy()
    global question_label, option_buttons, result_label, next_button, time_label
    time_label = tk.Label(root, text="Time Left: 300s", font=("Arial", 14))
    time_label.pack(pady=10)
    start_timer(duration)
    question_label = tk.Label(root, text="", font=("Arial", 16), wraplength=450, justify="center")
    question_label.pack(pady=20)
    option_buttons = []
    for i in range(4):
        button = tk.Button(
            root, text="", font=("Arial", 14), width=20, height=2, bg="lightgrey",
            command=lambda opt=i: check_answer(option_buttons[opt].cget("text"))
        )
        button.pack(pady=5)
        option_buttons.append(button)
    result_label = tk.Label(root, text="", font=("Arial", 14))
    result_label.pack(pady=10)
    next_button = tk.Button(root, text="Next", font=("Arial", 14), state="disabled", command=next_question)
    next_button.pack(pady=20)
    load_question()
# GUI setup
root = tk.Tk()
root.title("Quiz Application")
root.geometry("500x600")
start_screen()
root.mainloop()